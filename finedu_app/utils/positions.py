import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet



file_1 = load_workbook(
    '/Users/nuno/Django/Project-FinEdu/finedu_app/finedu_portfolio/portfolio_files/Investimentos.xlsx',
    data_only=True)
file_sheets_1 = file_1.sheetnames


def get_positions():
    year = 2022
    total_positions = []
    for sheet in file_sheets_1:
        fs = file_1[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        position = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if fs.cell(column, 4).fill.bgColor.index == dividend_color:
                    continue
                elif fs.cell(column, 4).fill.bgColor.index == close_position_color:
                    continue
                elif isinstance(cell_value, float):
                    position.append(cell_value)
                total = round(sum(position), 2)
        total_positions.append(total)
        # print(f'Posição em %s: €{total_position}' % ano)
        year += 1

    # print(f'Total: €{round(sum(total_position), 2)}')
    return total_positions


# List all the companies that are/have been a part of our investment portfolio
def companies():
    companies_set = set([])
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        for row in range(2, fs_count_row + 1):
            for column in range(2, 3):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                if not isinstance(cell_value, str):
                    continue
                cell_value = cell_value.strip()
                companies_set.add(cell_value)
    return list(companies_set)


def choose_companies(func):
    func = companies()
    char = input('Insira uma letra: ').title()

    if char == 'Gurufocus':
        return char

    def first_char(company_name):
        return company_name[0]

    temp_list_companies = sorted(func, key=first_char)
    grouped_list = [list(elements) for i, elements in groupby(temp_list_companies, first_char)]

    if char == '0':
        for company in temp_list_companies:
            print(company)

    for groups in grouped_list:
        i = 0
        for group in groups:
            if char == first_char(group[0]):
                print(group)
                i += 1


def companies_positions(cmp_name):
    i = 0
    # List with all the bought positions for each excel sheet (year)
    totals = []
    # Iterate through each excel sheet (year)
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        # Dict containing key=companyName, value=[positions]
        totals_bought = {}
        for row in range(2, fs_count_row + 1):
            year = file_sheets[i]
            for column in range(1, 4):
                # Define the cell
                cell = fs.cell(column=column, row=row)
                # Grab company name
                company = fs.cell(row, 2).value
                # Grab position cost
                position = fs.cell(row, 4).value
                # List all the positions bought for a company, per sheet (year)
                lst_positions = []
                # Exclude non str elements
                if not isinstance(company, str):
                    continue
                company = company.strip()

                # Exclude dividends
                if fs.cell(row, 4).fill.bgColor.index == dividend_color:
                    continue
                # Exclude sold positions
                if fs.cell(row, 4).fill.bgColor.index == close_position_color:
                    continue

                if company in totals_bought:
                    # Add position to existing list of positions
                    if not position in totals_bought[company]:
                        totals_bought[company].append(position)
                else:
                    # If dict is empty add key=companyName and value[positions] to dict
                    totals_bought[company] = lst_positions

        # Prints all the positions bought per year
        if cmp_name == 'ALL':
            print(f'Posições de ({year}): {totals_bought}')
            i += 1
        else:
            try:
                total = float([round(sum(totals_bought[cmp_name]), 2)][0])
                # Total positions bought per year
                print(
                    f'Posição comprada {cmp_name}: {totals_bought[cmp_name]} : Total = {total}')
                totals.append(total)
                i += 1

            except (KeyError, KeyboardInterrupt):
                print()

    if cmp_name == 'ALL':
        print()
        return 1
    else:
        sum_positions = round(sum(totals), 2)
        print(f'Total da posição: €{sum_positions}')
        print('------------------------------------')
        return sum_positions


def sold_positions(cmp_name):
    year = 2022
    totals = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        totals_sold = {}
        cnt = 1
        for row in range(2, fs_count_row + 1):
            for column in range(2, 4):
                cell = fs.cell(column=column, row=row)
                company = fs.cell(row, 2).value
                position = fs.cell(row, 4).value
                lst_positions = []
                if not isinstance(company, str):
                    continue
                company = company.strip()
                if fs.cell(row, 4).fill.bgColor.index == dividend_color:
                    continue
                if fs.cell(row, 4).fill.bgColor.index == close_position_color:
                    if company in totals_sold:
                        if not position in totals_sold[company]:
                            totals_sold[company].append(position)
                    else:
                        totals_sold[company] = lst_positions

        # Prints all the positions sold per year
        if cmp_name == 'ALL':
            print(f'Posições de %s: {totals_sold}' % year)
            year += 1
        else:
            try:
                total = float([round(sum(totals_sold[cmp_name]), 2)][0])
                while total == 0:
                    cnt += 1
                    continue
                else:
                    print(f'Posição vendida {cmp_name} (%s): {totals_sold[cmp_name]} : Total = {total}' % (
                            year + cnt))
                    totals.append(total)
                    year += 1
            except (KeyError, KeyboardInterrupt):
                print()

    if cmp_name == 'ALL':
        return 1
    else:
        sum_sold_positions = round(sum(totals), 2)
        print(f'Posição final de venda: €{sum_sold_positions * -1}')
        return sum_sold_positions


# Check for profit or loss when position was sold
def profit_loss(cmp_name, bought, sold):
    if bought and sold == 1:
        print('Sem dados de lucro ou perda')
    else:
        sold_value = abs(sold)
        if sold_value > bought:
            profit = sold_value - bought
            percentage = sold_value * 100 / bought
            percentage_profit = abs(100 - percentage)
            print(f'Vendido com lucro de €{round(profit, 2)} : +{round(percentage_profit, 2)}%')
        elif sold_value == 0:
            print('Ainda não vendeu a posição na empresa.')
            return 0
        else:
            loss = bought - sold_value
            percentage = sold_value * 100 / bought
            percentage_loss = abs(100 - percentage)
            print(f'Vendido com perda de €{round(loss, 2)} : -{round(percentage_loss, 2)}%')
        return cmp_name

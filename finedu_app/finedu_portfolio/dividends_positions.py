"""
Python openpyxl
Python Seaborn
Python Plotly

"""

from io import BytesIO
from typing import Dict, List

import openpyxl
import pandas as pd
import plotly.express as px
from finedu_portfolio.company_info import get_current_positions
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DIVIDEND_COLOR = 'FFFFFF00'  # yellow
CLOSE_POSITION_COLOR = 'FFFF0000'  # red
NR_OF_SHARES_COLOR = '#FFFF00'  # orange


file = load_workbook(
    '/app/portfolio_files/Investimentos.xlsx',
    data_only=True)
file_sheets = file.sheetnames


def get_dividends():
    """
    """
    year = 2022
    total_dividends = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        dividends = []
        for row in range(2, fs_count_row + 1):
            for column in range(4, 5):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if bgColor == DIVIDEND_COLOR:
                    dividends.append(cell_value)
                total = sum(dividends)
        total_dividends.append(total)
        # print(f'Dividendos de %s: €{total:.2f}' % ano)
        year += 1
    # print(f'Total: €{round(sum(total_dividends), 2)}')
    # print('------------------------------------')
    return total_dividends


def dividends_per_year(func, year):
    dividends_dict = {}
    for i, dividend in enumerate(func):
        dividends_dict[i + year] = round(dividend, 2)
    return dividends_dict


def dividends_growth(dividends):
    dividends = dividends_per_year(get_dividends(), 2022)
    previous_year_amount = None
    for year, amount in dividends.items():
        yoy = None
        if previous_year_amount is not None:
            yoy = (amount - previous_year_amount) / previous_year_amount * 100
        dividends[year] = {'amount': amount, 'YoY': yoy}
        previous_year_amount = amount
    return dividends


# div_growth = dividends_growth(dividends_per_year(get_dividends(), 2022))
# print(div_growth)

def get_dividends_growth():
    """
    """
    div_growth = dividends_growth(dividends_per_year(get_dividends(), 2022))
    growth_yoy = []
    for year in div_growth.items():
        growth_yoy.append(year[1]['YoY'])
    return growth_yoy


def dividend_data_graph():
    dividends = dividends_per_year(get_dividends(), 2022)
    x_graph = []
    y_graph = []
    for item in dividends.items():
        x_graph.append(item[0])
        y_graph.append(item[1])

    formated_y_graph = [float(y) for y in y_graph]

    return x_graph, formated_y_graph


total_dividends_growth_yoy = dividends_growth(
    dividends=dividends_per_year(get_dividends(), 2022))


def dividend_and_position():
    """
    """
    dividend_and_position = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        for row in range(2, fs_count_row + 1):
            temp = []
            if fs.cell(column=4, row=row).fill.bgColor.index == DIVIDEND_COLOR:
                for col in range(2, 5):
                    cell_value = fs.cell(row=row, column=col).value
                    temp.append(cell_value)
                dividend_and_position.append(temp)
                # Print with a space after each value
                # print(cell_value, end=" ")
            # print()  # Print a new line after each row
    return dividend_and_position


div_and_pos = dividend_and_position()
# print(div_and_pos)
# print()


def dividends_per_position(func: List, company):
    """
    """
    # Find the index of the dividend column (assuming the third column)
    company_dividends = {}
    dividend_index = 2
    dividends = [item for item in func if item[0] == company]
    sum_dividends = sum(item[dividend_index] for item in dividends)
    company_dividends[company] = round(sum_dividends, 2)

    return company_dividends


def get_previous_portfolio_positions():
    """
    """
    current_portfolio = get_current_positions()
    previous_portolio_positions = []

    for item in div_and_pos:
        if not item[0] in current_portfolio:
            previous_portolio_positions.append(item[0])

    return list(set(previous_portolio_positions))

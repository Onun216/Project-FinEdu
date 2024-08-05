import os
from itertools import groupby

import import_export
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ENTER_POSITION_COLOR = 'FF00FF00'  # green


file = load_workbook(
    '/app/portfolio_files/Portfolio.xlsx',
    data_only=True)
file_sheets = file.sheetnames


def get_current_positions():
    """
    """
    names = []
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_column = fs.max_column
        for row in range(2, fs_count_row + 1):
            for column in range(1, 2):
                cell = fs.cell(column=column, row=row)
                cell_value = fs.cell(row, column).value
                bgColor = cell.fill.bgColor.index
                if bgColor == ENTER_POSITION_COLOR:
                    names.append(cell_value)
        return names


def get_current_positions_details():
    """
    """
    details = {}
    for sheet in file_sheets:
        fs = file[sheet]
        fs_count_row = fs.max_row
        fs_count_row = fs.max_column
        for row in range(2, fs_count_row + 1):
            temp_details_list = []
            for column in range(2, 5):
                cell = fs.cell(column=column, row=row)
                if fs.cell(column=1, row=row).fill.bgColor.index == ENTER_POSITION_COLOR:
                    cell = fs.cell(row, column).value
                    temp_details_list.append(cell)
                    details[fs.cell(row, column=1).value] = temp_details_list
    return details


current_positions = get_current_positions()
current_positions_details = get_current_positions_details()
# print(current_positions_details)


def locations(details):
    """
    """
    latitude_list = []
    longitude_list = []
    for company in details:
        company_info = details[company]
        latitude = company_info[1]
        longitude = company_info[2]
        latitude_list.append(latitude)
        longitude_list.append(longitude)
    return latitude_list, longitude_list
